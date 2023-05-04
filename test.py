import openpyxl

# otevřít excelový soubor
workbook = openpyxl.load_workbook('nazev_souboru.xlsx')

# vybrat konkrétní list
worksheet = workbook['nazev_listu']

# projít všechny řádky a najít hledané řádky
for row in worksheet.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and 'ISIN' in cell.value:
            # najít buňku vedle ISIN s číselnou hodnotou
            next_cell = cell.offset(column=1)
            if isinstance(next_cell.value, (int, float)):
                print(cell.value, next_cell.value)





#---- ověří zda jde o ISIN na základě syntaxe kódu

import re
import openpyxl

# otevřít excelový soubor
workbook = openpyxl.load_workbook('nazev_souboru.xlsx')

# vybrat konkrétní list
worksheet = workbook['nazev_listu']

# definovat regulární výraz pro ISIN kódy
isin_regex = re.compile(r'\b[A-Z]{2}[A-Z0-9]{9}[0-9]\b')

# projít všechny řádky a najít hledané řádky
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value is not None and isin_regex.match(str(cell.value)):
            print(cell.value)
