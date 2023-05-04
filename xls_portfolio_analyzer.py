import openpyxl

# otevřít excelový soubor
workbook = openpyxl.load_workbook('nazev_souboru.xlsx')

# vybrat konkrétní list
worksheet = workbook['nazev_listu']

# získat hodnoty z buněk
value1 = worksheet['A1'].value
value2 = worksheet['B2'].value

# otevřít textový soubor
with open('nazev_textoveho_souboru.txt', 'r') as f:
    # přečíst řádek s hlavičkou
    header = f.readline().strip()
    # přečíst řádky s hodnotami
    rows = [line.strip().split('\t') for line in f.readlines()]

# vytvořit nový excelový soubor
new_workbook = openpyxl.Workbook()

# vytvořit nový list
new_worksheet = new_workbook.active
new_worksheet.title = 'novy_list'

# nastavit hlavičku na prvním řádku
header_row = header.split('\t')
for i, header_val in enumerate(header_row):
    new_worksheet.cell(row=1, column=i+1, value=header_val)

# projít řádky s hodnotami a vypočítat nové hodnoty
for row_num, row in enumerate(rows):
    # vypočítat novou hodnotu
    new_val = row[0] + row[1]
    # zapsat novou hodnotu do nového souboru
    new_worksheet.cell(row=row_num+2, column=1, value=new_val)

# uložit nový excelový soubor
new_workbook.save('novy_nazev_souboru.xlsx')
